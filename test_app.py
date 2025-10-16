#!/usr/bin/env python3
"""
Tests for the Data Sorter Application
"""

import pytest
import os
import tempfile
from unittest.mock import Mock, patch
from app import DataSorterApp
from openpyxl import load_workbook


class TestDataSorterApp:
    """Test suite for DataSorterApp class."""
    
    @pytest.fixture
    def app(self):
        """Create a DataSorterApp instance for testing without GUI."""
        with patch('app.tk.Tk'):
            root = Mock()
            app = DataSorterApp.__new__(DataSorterApp)
            app.root = root
            return app
    
    def test_parse_records_single_record(self, app):
        """Test parsing a single record."""
        text = """Name: John Doe
CO-OP NAME: Alpha Co-op
Member ID: 12345"""
        
        records = app.parse_records(text)
        
        assert len(records) == 1
        assert records[0]['Name'] == 'John Doe'
        assert records[0]['CO-OP NAME'] == 'Alpha Co-op'
        assert records[0]['Member ID'] == '12345'
    
    def test_parse_records_multiple_records(self, app):
        """Test parsing multiple records separated by blank lines."""
        text = """Name: John Doe
CO-OP NAME: Alpha Co-op
Member ID: 12345

Name: Jane Smith
CO-OP NAME: Beta Co-op
Member ID: 67890"""
        
        records = app.parse_records(text)
        
        assert len(records) == 2
        assert records[0]['Name'] == 'John Doe'
        assert records[1]['Name'] == 'Jane Smith'
    
    def test_parse_records_empty_input(self, app):
        """Test parsing empty input."""
        text = ""
        
        records = app.parse_records(text)
        
        assert len(records) == 0
    
    def test_parse_records_whitespace_only(self, app):
        """Test parsing whitespace-only input."""
        text = "   \n\n   \n"
        
        records = app.parse_records(text)
        
        assert len(records) == 0
    
    def test_parse_records_with_trailing_blank_lines(self, app):
        """Test parsing records with trailing blank lines."""
        text = """Name: John Doe
CO-OP NAME: Alpha Co-op
Member ID: 12345

"""
        
        records = app.parse_records(text)
        
        assert len(records) == 1
        assert records[0]['Name'] == 'John Doe'
    
    def test_parse_records_without_trailing_blank_line(self, app):
        """Test parsing records without trailing blank line."""
        text = """Name: John Doe
CO-OP NAME: Alpha Co-op
Member ID: 12345"""
        
        records = app.parse_records(text)
        
        assert len(records) == 1
        assert records[0]['Name'] == 'John Doe'
    
    def test_parse_records_with_colon_in_value(self, app):
        """Test parsing records where value contains colon."""
        text = """Email: john@example.com
URL: https://example.com:8080
CO-OP NAME: Alpha Co-op"""
        
        records = app.parse_records(text)
        
        assert len(records) == 1
        assert records[0]['Email'] == 'john@example.com'
        assert records[0]['URL'] == 'https://example.com:8080'
    
    def test_parse_records_with_empty_value(self, app):
        """Test parsing records with empty values."""
        text = """Name: John Doe
Email:
CO-OP NAME: Alpha Co-op"""
        
        records = app.parse_records(text)
        
        assert len(records) == 1
        assert records[0]['Email'] == ''
    
    def test_parse_records_multiple_blank_lines(self, app):
        """Test parsing records separated by multiple blank lines."""
        text = """Name: John Doe
CO-OP NAME: Alpha Co-op


Name: Jane Smith
CO-OP NAME: Beta Co-op"""
        
        records = app.parse_records(text)
        
        assert len(records) == 2
        assert records[0]['Name'] == 'John Doe'
        assert records[1]['Name'] == 'Jane Smith'
    
    def test_group_by_coop_name_single_coop(self, app):
        """Test grouping records from a single co-op."""
        records = [
            {'Name': 'John Doe', 'CO-OP NAME': 'Alpha Co-op'},
            {'Name': 'Jane Smith', 'CO-OP NAME': 'Alpha Co-op'}
        ]
        
        grouped = app.group_by_coop_name(records)
        
        assert len(grouped) == 1
        assert 'Alpha Co-op' in grouped
        assert len(grouped['Alpha Co-op']) == 2
    
    def test_group_by_coop_name_multiple_coops(self, app):
        """Test grouping records from multiple co-ops."""
        records = [
            {'Name': 'John Doe', 'CO-OP NAME': 'Alpha Co-op'},
            {'Name': 'Jane Smith', 'CO-OP NAME': 'Beta Co-op'},
            {'Name': 'Bob Johnson', 'CO-OP NAME': 'Alpha Co-op'}
        ]
        
        grouped = app.group_by_coop_name(records)
        
        assert len(grouped) == 2
        assert 'Alpha Co-op' in grouped
        assert 'Beta Co-op' in grouped
        assert len(grouped['Alpha Co-op']) == 2
        assert len(grouped['Beta Co-op']) == 1
    
    def test_group_by_coop_name_missing_coop_field(self, app):
        """Test grouping records without CO-OP NAME field."""
        records = [
            {'Name': 'John Doe'},
            {'Name': 'Jane Smith', 'CO-OP NAME': 'Alpha Co-op'}
        ]
        
        grouped = app.group_by_coop_name(records)
        
        assert len(grouped) == 2
        assert 'Unknown' in grouped
        assert 'Alpha Co-op' in grouped
        assert len(grouped['Unknown']) == 1
    
    def test_group_by_coop_name_empty_records(self, app):
        """Test grouping empty records list."""
        records = []
        
        grouped = app.group_by_coop_name(records)
        
        assert len(grouped) == 0
    
    def test_create_excel_file_single_sheet(self, app):
        """Test creating Excel file with a single sheet."""
        grouped_data = {
            'Alpha Co-op': [
                {'Name': 'John Doe', 'Member ID': '12345'},
                {'Name': 'Jane Smith', 'Member ID': '67890'}
            ]
        }
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filepath = tmp.name
        
        try:
            app.create_excel_file(grouped_data, filepath)
            
            # Verify the file was created
            assert os.path.exists(filepath)
            
            # Load and verify content
            wb = load_workbook(filepath)
            assert 'Alpha Co-op' in wb.sheetnames
            
            sheet = wb['Alpha Co-op']
            assert sheet.cell(row=1, column=1).value == 'Name'
            assert sheet.cell(row=1, column=2).value == 'Member ID'
            assert sheet.cell(row=2, column=1).value == 'John Doe'
            assert sheet.cell(row=2, column=2).value == '12345'
            assert sheet.cell(row=3, column=1).value == 'Jane Smith'
            assert sheet.cell(row=3, column=2).value == '67890'
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
    
    def test_create_excel_file_multiple_sheets(self, app):
        """Test creating Excel file with multiple sheets."""
        grouped_data = {
            'Alpha Co-op': [
                {'Name': 'John Doe', 'Member ID': '12345'}
            ],
            'Beta Co-op': [
                {'Name': 'Bob Johnson', 'Member ID': '11111'}
            ]
        }
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filepath = tmp.name
        
        try:
            app.create_excel_file(grouped_data, filepath)
            
            wb = load_workbook(filepath)
            assert 'Alpha Co-op' in wb.sheetnames
            assert 'Beta Co-op' in wb.sheetnames
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
    
    def test_create_excel_file_long_sheet_name(self, app):
        """Test creating Excel file with long co-op name (sheet name limit is 31 chars)."""
        long_name = 'A' * 40  # 40 characters, exceeds limit
        grouped_data = {
            long_name: [
                {'Name': 'John Doe', 'Member ID': '12345'}
            ]
        }
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filepath = tmp.name
        
        try:
            app.create_excel_file(grouped_data, filepath)
            
            wb = load_workbook(filepath)
            # Sheet name should be truncated to 31 characters
            assert long_name[:31] in wb.sheetnames
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
    
    def test_create_excel_file_duplicate_sheet_names(self, app):
        """Test handling duplicate sheet names."""
        # This would create duplicate sheet names after truncation
        name1 = 'A' * 31 + '1'
        name2 = 'A' * 31 + '2'
        
        grouped_data = {
            name1: [{'Name': 'John Doe'}],
            name2: [{'Name': 'Jane Smith'}]
        }
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filepath = tmp.name
        
        try:
            app.create_excel_file(grouped_data, filepath)
            
            wb = load_workbook(filepath)
            # Should have 2 sheets with different names
            assert len(wb.sheetnames) == 2
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
    
    def test_create_excel_file_varying_fields(self, app):
        """Test creating Excel with records having different fields."""
        grouped_data = {
            'Alpha Co-op': [
                {'Name': 'John Doe', 'Email': 'john@example.com'},
                {'Name': 'Jane Smith', 'Phone': '555-1234'}
            ]
        }
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filepath = tmp.name
        
        try:
            app.create_excel_file(grouped_data, filepath)
            
            wb = load_workbook(filepath)
            sheet = wb['Alpha Co-op']
            
            # Should have all unique fields in headers
            headers = [sheet.cell(row=1, column=i).value for i in range(1, 4)]
            assert 'Name' in headers
            assert 'Email' in headers
            assert 'Phone' in headers
            
            # Second record should have empty cell for Email
            row2_data = [sheet.cell(row=3, column=i).value for i in range(1, 4)]
            assert 'Jane Smith' in row2_data
            assert '555-1234' in row2_data
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
    
    def test_create_excel_file_column_width_adjustment(self, app):
        """Test that column widths are properly adjusted."""
        grouped_data = {
            'Alpha Co-op': [
                {'Name': 'John', 'Description': 'A' * 60}  # Very long text
            ]
        }
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filepath = tmp.name
        
        try:
            app.create_excel_file(grouped_data, filepath)
            
            wb = load_workbook(filepath)
            sheet = wb['Alpha Co-op']
            
            # Column width for 'Description' should be capped at 50
            from openpyxl.utils import get_column_letter
            desc_col = get_column_letter(2)  # Assuming Description is second column
            assert sheet.column_dimensions[desc_col].width <= 50
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
    
    def test_integration_parse_group_export(self, app):
        """Integration test: parse -> group -> export."""
        text = """Name: John Doe
CO-OP NAME: Alpha Co-op
Member ID: 12345

Name: Jane Smith
CO-OP NAME: Alpha Co-op
Member ID: 67890

Name: Bob Johnson
CO-OP NAME: Beta Co-op
Member ID: 11111"""
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filepath = tmp.name
        
        try:
            # Parse
            records = app.parse_records(text)
            assert len(records) == 3
            
            # Group
            grouped = app.group_by_coop_name(records)
            assert len(grouped) == 2
            
            # Export
            app.create_excel_file(grouped, filepath)
            
            # Verify
            wb = load_workbook(filepath)
            assert 'Alpha Co-op' in wb.sheetnames
            assert 'Beta Co-op' in wb.sheetnames
            
            alpha_sheet = wb['Alpha Co-op']
            beta_sheet = wb['Beta Co-op']
            
            # Alpha Co-op should have 2 records (+ 1 header = 3 rows)
            assert alpha_sheet.max_row == 3
            # Beta Co-op should have 1 record (+ 1 header = 2 rows)
            assert beta_sheet.max_row == 2
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
