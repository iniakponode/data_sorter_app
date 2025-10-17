#!/usr/bin/env python3
"""
SmartDataExtractor Application
A Tkinter-based desktop application that intelligently parses unstructured text data,
extracts key-value pairs, handles orphaned values, and exports to structured Excel files.

Developer: Iniakpokeikiye Peter Thompson, CTO
Company: Ungozu and Son's Enterprises Limited
Version: 2.9.0
Copyright (c) 2025 Ungozu and Son's Enterprises Limited
"""

import tkinter as tk
from tkinter import scrolledtext, messagebox, filedialog, ttk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from collections import defaultdict
import re
import os
try:
    import docx  # For Word document processing
except ImportError:
    docx = None
try:
    import PyPDF2  # For PDF processing
except ImportError:
    PyPDF2 = None


class SmartDataExtractorApp:
    """Main application class for the SmartDataExtractor."""
    
    def __init__(self, root):
        """Initialize the application UI."""
        self.root = root
        
        # Standard column configuration
        self.standard_columns = [
            "S/N",
            "NAME OF COOPERATIVE", 
            "CEO NAME",
            "PHONE No.",
            "BANK NAME",
            "ACNT. No.",
            "SEX"
        ]
        
        # User-configurable columns (starts with standard)
        self.user_columns = self.standard_columns.copy()
        
        # Record boundary configuration
        self.record_start_field = "NAME OF COOPERATIVE"  # Default start field
        self.record_end_field = "SEX"  # Default end field
        
        # Create UI elements only if root is provided
        if self.root:
            self.root.title("SmartDataExtractor v2.9.0 - by Ungozu and Son's Enterprises")
            self.root.geometry("900x700")
            self.create_menu()
            self.create_widgets()
    
    def create_menu(self):
        """Create the application menu bar."""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
    
    def show_about(self):
        """Show application about dialog."""
        about_text = """SmartDataExtractor v2.9.0

Intelligent Data Processing Application

Developer: Iniakpokeikiye Peter Thompson
Title: Chief Technology Officer (CTO)
Company: Ungozu and Son's Enterprises Limited

Features:
• Intelligent parsing of unstructured text data
• Smart orphaned value detection and assignment
• Advanced pattern recognition with 7-priority system
• Export to structured Excel files
• Support for Word and PDF file uploads

Copyright © 2025 Ungozu and Son's Enterprises Limited
All rights reserved."""
        
        messagebox.showinfo("About SmartDataExtractor", about_text)

    def create_widgets(self):
        """Create and layout UI widgets with enhanced features."""
        # Create main notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Data Input Tab
        input_frame = ttk.Frame(notebook)
        notebook.add(input_frame, text="Data Input")
        
        # Instructions
        instructions = tk.Label(
            input_frame,
            text="Enter your data below or upload a file (Word/PDF supported)\n" +
                 "Data will be automatically parsed and organized into standard columns",
            font=("Arial", 10),
            pady=10,
            justify=tk.LEFT
        )
        instructions.pack()
        
        # File upload frame
        file_frame = tk.Frame(input_frame)
        file_frame.pack(pady=5)
        
        tk.Button(
            file_frame,
            text="Upload File (Word/PDF)",
            command=self.upload_file,
            font=("Arial", 10),
            bg="#2196F3",
            fg="white",
            padx=15,
            pady=5
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            file_frame,
            text="Clear Text",
            command=self.clear_text,
            font=("Arial", 10),
            padx=15,
            pady=5
        ).pack(side=tk.LEFT, padx=5)
        
        # Text area for input
        self.text_area = scrolledtext.ScrolledText(
            input_frame,
            width=100,
            height=20,
            font=("Courier", 9),
            wrap=tk.WORD
        )
        self.text_area.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        
        # Column Configuration Tab
        config_frame = ttk.Frame(notebook)
        notebook.add(config_frame, text="Column Configuration")
        
        # Column management section
        col_mgmt_frame = tk.Frame(config_frame)
        col_mgmt_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Label(
            col_mgmt_frame,
            text="Manage Output Columns",
            font=("Arial", 12, "bold")
        ).pack(pady=(0, 10))
        
        # Current columns listbox
        list_frame = tk.Frame(col_mgmt_frame)
        list_frame.pack(fill="both", expand=True)
        
        tk.Label(list_frame, text="Current Columns:", font=("Arial", 10)).pack(anchor="w")
        
        # Frame for listbox and scrollbar
        listbox_frame = tk.Frame(list_frame)
        listbox_frame.pack(fill="both", expand=True, pady=5)
        
        self.columns_listbox = tk.Listbox(listbox_frame, height=10, font=("Arial", 10))
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical")
        self.columns_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.columns_listbox.yview)
        
        self.columns_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Populate initial columns
        self.refresh_columns_list()
        
        # Column management buttons
        btn_frame = tk.Frame(col_mgmt_frame)
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="Add Column",
            command=self.add_column,
            font=("Arial", 10),
            bg="#4CAF50",
            fg="white",
            padx=15
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Delete Selected",
            command=self.delete_column,
            font=("Arial", 10),
            bg="#f44336",
            fg="white",
            padx=15
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="Reset to Standard",
            command=self.reset_columns,
            font=("Arial", 10),
            padx=15
        ).pack(side=tk.LEFT, padx=5)
        
        # Record Boundary Configuration Section
        boundary_frame = tk.LabelFrame(col_mgmt_frame, text="Record Boundary Configuration", font=("Arial", 11, "bold"))
        boundary_frame.pack(fill="x", pady=(20, 0))
        
        tk.Label(
            boundary_frame,
            text="Define which fields mark the start and end of a complete record:",
            font=("Arial", 10),
            justify=tk.LEFT
        ).pack(anchor="w", padx=10, pady=(5, 0))
        
        tk.Label(
            boundary_frame,
            text="This helps properly parse multi-line records and group related information.",
            font=("Arial", 9),
            fg="gray",
            justify=tk.LEFT
        ).pack(anchor="w", padx=10, pady=(0, 10))
        
        # Start field selection
        start_frame = tk.Frame(boundary_frame)
        start_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(start_frame, text="Record starts with:", font=("Arial", 10)).pack(side=tk.LEFT)
        
        self.start_field_var = tk.StringVar(value=self.record_start_field)
        self.start_field_combo = ttk.Combobox(
            start_frame,
            textvariable=self.start_field_var,
            values=self.user_columns,
            state="readonly",
            width=25
        )
        self.start_field_combo.pack(side=tk.LEFT, padx=(10, 0))
        self.start_field_combo.bind('<<ComboboxSelected>>', self.update_record_start_field)
        
        # End field selection
        end_frame = tk.Frame(boundary_frame)
        end_frame.pack(fill="x", padx=10, pady=(5, 15))
        
        tk.Label(end_frame, text="Record ends with:", font=("Arial", 10)).pack(side=tk.LEFT)
        
        self.end_field_var = tk.StringVar(value=self.record_end_field)
        self.end_field_combo = ttk.Combobox(
            end_frame,
            textvariable=self.end_field_var,
            values=self.user_columns,
            state="readonly",
            width=25
        )
        self.end_field_combo.pack(side=tk.LEFT, padx=(10, 0))
        self.end_field_combo.bind('<<ComboboxSelected>>', self.update_record_end_field)

        # Processing Options Tab
        options_frame = ttk.Frame(notebook)
        notebook.add(options_frame, text="Processing Options")
        
        # Sheet organization option
        sheet_frame = tk.LabelFrame(options_frame, text="Output Format", font=("Arial", 11, "bold"))
        sheet_frame.pack(fill="x", padx=10, pady=10)
        
        self.sheet_mode_var = tk.StringVar(value="separate")
        
        tk.Radiobutton(
            sheet_frame,
            text="Separate sheets by Cooperative Name (recommended)",
            variable=self.sheet_mode_var,
            value="separate",
            font=("Arial", 10)
        ).pack(anchor="w", padx=10, pady=5)
        
        tk.Radiobutton(
            sheet_frame,
            text="All records in one sheet",
            variable=self.sheet_mode_var,
            value="single",
            font=("Arial", 10)
        ).pack(anchor="w", padx=10, pady=5)
        
        # Process button (main action)
        process_frame = tk.Frame(self.root)
        process_frame.pack(pady=10)
        
        self.process_button = tk.Button(
            process_frame,
            text="Process Data and Export to Excel",
            command=self.process_data,
            font=("Arial", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            padx=30,
            pady=15
        )
        self.process_button.pack()
    
    def upload_file(self):
        """Handle file upload and extract text content."""
        file_path = filedialog.askopenfilename(
            title="Select file to upload",
            filetypes=[
                ("Word documents", "*.docx"),
                ("PDF files", "*.pdf"),
                ("Text files", "*.txt"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            text_content = ""
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.docx':
                text_content = self.extract_text_from_word(file_path)
            elif file_ext == '.pdf':
                text_content = self.extract_text_from_pdf(file_path)
            elif file_ext == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    text_content = f.read()
            else:
                messagebox.showerror("Error", "Unsupported file format. Please use .docx, .pdf, or .txt files.")
                return
            
            # Insert content into text area
            self.text_area.delete("1.0", tk.END)
            self.text_area.insert("1.0", text_content)
            
            messagebox.showinfo("Success", f"File content loaded successfully!\nExtracted {len(text_content)} characters.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read file: {str(e)}")
    
    def extract_text_from_word(self, file_path):
        """Extract text from Word document."""
        if docx is None:
            raise ImportError("python-docx library not installed. Cannot read Word files.")
        
        doc = docx.Document(file_path)
        text_content = []
        
        for paragraph in doc.paragraphs:
            text_content.append(paragraph.text)
        
        return '\n'.join(text_content)
    
    def extract_text_from_pdf(self, file_path):
        """Extract text from PDF file."""
        if PyPDF2 is None:
            raise ImportError("PyPDF2 library not installed. Cannot read PDF files.")
        
        text_content = []
        
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page in pdf_reader.pages:
                text_content.append(page.extract_text())
        
        return '\n'.join(text_content)
    
    def clear_text(self):
        """Clear the text area."""
        self.text_area.delete("1.0", tk.END)
    
    def refresh_columns_list(self):
        """Refresh the columns listbox with current user columns."""
        self.columns_listbox.delete(0, tk.END)
        for i, column in enumerate(self.user_columns):
            display_text = f"{i+1}. {column}"
            self.columns_listbox.insert(tk.END, display_text)
        # Update record boundary comboboxes when columns change
        self.update_boundary_comboboxes()
    
    def add_column(self):
        """Add a new column to the configuration."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Add New Column")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Enter column name:", font=("Arial", 10)).pack(pady=10)
        
        entry = tk.Entry(dialog, font=("Arial", 10), width=30)
        entry.pack(pady=5)
        entry.focus()
        
        def add_action():
            column_name = entry.get().strip()
            if column_name:
                if column_name not in self.user_columns:
                    self.user_columns.append(column_name)
                    self.refresh_columns_list()
                    dialog.destroy()
                else:
                    messagebox.showwarning("Warning", "Column already exists!")
            else:
                messagebox.showwarning("Warning", "Please enter a column name!")
        
        def cancel_action():
            dialog.destroy()
        
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="Add", command=add_action, bg="#4CAF50", fg="white").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=cancel_action).pack(side=tk.LEFT, padx=5)
        
        # Bind Enter key to add action
        entry.bind('<Return>', lambda e: add_action())
    
    def delete_column(self):
        """Delete selected column from configuration."""
        selection = self.columns_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a column to delete!")
            return
        
        index = selection[0]
        column_name = self.user_columns[index]
        
        # Prevent deletion of S/N column
        if column_name == "S/N":
            messagebox.showwarning("Warning", "Cannot delete S/N column - it's automatically generated!")
            return
        
        if messagebox.askyesno("Confirm", f"Delete column '{column_name}'?"):
            del self.user_columns[index]
            self.refresh_columns_list()
    
    def reset_columns(self):
        """Reset columns to standard configuration."""
        if messagebox.askyesno("Confirm", "Reset to standard columns? This will remove any custom columns."):
            self.user_columns = self.standard_columns.copy()
            self.refresh_columns_list()
            # Update record boundary comboboxes
            self.update_boundary_comboboxes()
    
    def update_record_start_field(self, event=None):
        """Update the record start field when combobox selection changes."""
        self.record_start_field = self.start_field_var.get()
    
    def update_record_end_field(self, event=None):
        """Update the record end field when combobox selection changes."""
        self.record_end_field = self.end_field_var.get()
    
    def update_boundary_comboboxes(self):
        """Update the values in record boundary comboboxes when columns change."""
        if hasattr(self, 'start_field_combo') and hasattr(self, 'end_field_combo'):
            self.start_field_combo['values'] = self.user_columns
            self.end_field_combo['values'] = self.user_columns
    
    def is_intelligent_record_boundary(self, line, current_block, all_lines, current_index):
        """
        Intelligently detect record boundaries based on content patterns.
        
        Args:
            line (str): Current line being processed
            current_block (list): Current record block being built
            all_lines (list): All lines in the text
            current_index (int): Index of current line
            
        Returns:
            bool: True if this line should start a new record
        """
        if not current_block:
            return False
            
        line_upper = line.upper().strip()
        
        # Look for strong indicators of a new record starting
        cooperative_indicators = ['COOPERATIVE NAME', 'CO-OP NAME', 'COOP NAME', 'ORGANIZATION NAME']
        
        # If current line looks like a cooperative name field and we have a substantial current block
        if any(indicator in line_upper for indicator in cooperative_indicators):
            if len(current_block) >= 3:  # Minimum lines for a complete record
                return True
        
        # Check if this looks like the start of a new record pattern
        # by looking for repeating field patterns
        if len(current_block) >= 4:  # Only if we have a substantial record already
            # Check if we see a pattern that suggests new record
            key, value = self.extract_key_value_from_line(line)
            if key:
                key_normalized = self.normalize_key_name(key)
                # If this field already exists in current block, might be new record
                block_text = '\n'.join(current_block)
                if key_normalized in ['CO-OP NAME', 'COOPERATIVE', 'NAME'] and key_normalized not in block_text.upper():
                    # This is a major field that should start a new record
                    return True
        
        return False

    def merge_orphaned_blocks(self, record_blocks):
        """
        Merge blocks that contain only orphaned values with preceding complete records.
        
        Args:
            record_blocks (list): List of record blocks
            
        Returns:
            list: Merged record blocks
        """
        if not record_blocks:
            return []
        
        merged_blocks = []
        
        for i, block in enumerate(record_blocks):
            # Check if this block looks like orphaned values
            is_orphaned_block = self.is_likely_orphaned_block(block)
            
            if is_orphaned_block and merged_blocks:
                # Merge with the previous block
                merged_blocks[-1].extend(block)
            else:
                # Start a new block
                merged_blocks.append(block[:])  # Copy the block
        
        return merged_blocks
    
    def is_likely_orphaned_block(self, block):
        """
        Determine if a block likely contains orphaned values that should be merged.
        
        Args:
            block (list): Lines in the block
            
        Returns:
            bool: True if block should be merged with previous record
        """
        if not block:
            return False
        
        # A block should be merged if:
        # 1. It starts with an orphaned value (no key-value pair), OR
        # 2. It's short and doesn't contain major record start indicators
        
        first_line = block[0].strip()
        first_key, first_value = self.extract_key_value_from_line(first_line)
        
        # If first line is not a key-value pair, check if it looks like an orphaned value
        if not first_key:
            first_line_upper = first_line.upper()
            
            # Check if first line is a standalone value that should belong to previous record
            orphaned_indicators = [
                # Bank names
                any(bank in first_line_upper for bank in ['BANK', 'UBA', 'GTB', 'ACCESS', 'ZENITH', 'FIRST', 'POLARIS', 'FCMB', 'FIDELITY', 'STERLING']),
                # Gender values
                first_line_upper in ['MALE', 'FEMALE', 'M', 'F'],
                # Phone numbers (standalone digits)
                first_line.isdigit() and len(first_line) == 11,
                # Account numbers (standalone digits)
                first_line.isdigit() and len(first_line) >= 8
            ]
            
            if any(orphaned_indicators):
                return True
        
        # Additional check: if block is short and doesn't have major record indicators
        if len(block) <= 3:
            block_text = ' '.join(block).upper()
            major_indicators = ['COOPERATIVE NAME', 'CO-OP NAME', 'ORGANIZATION NAME']
            
            has_major_field = any(indicator in block_text for indicator in major_indicators)
            
            if not has_major_field:
                # Check if it contains typical orphaned patterns
                orphaned_count = 0
                for line in block:
                    line_upper = line.upper().strip()
                    key, value = self.extract_key_value_from_line(line)
                    
                    # Count explicit key-value pairs as non-orphaned
                    if key:
                        continue
                    
                    # Count standalone values that look orphaned
                    if (any(bank in line_upper for bank in ['BANK', 'UBA', 'GTB', 'ACCESS', 'ZENITH', 'FIRST']) or
                        line_upper in ['MALE', 'FEMALE', 'M', 'F'] or
                        (line.strip().isdigit() and len(line.strip()) >= 8)):
                        orphaned_count += 1
                
                # If we have orphaned values and no major fields, merge it
                return orphaned_count > 0
        
        return False

    def is_noise_line(self, line):
        """
        Determine if a line is noise/extraneous text that should be filtered out.
        
        Args:
            line (str): Line to check
            
        Returns:
            bool: True if the line is noise, False if it's potentially valid data
        """
        line_upper = line.upper()
        
        # First check if it's a valid key-value pair - if so, it's NOT noise
        if self.is_valid_key_value_pair(line):
            return False
        
        # Common noise patterns (only for non-key-value lines)
        noise_patterns = [
            # Headers and titles
            'PERSONAL DATA', 'SOCIETY', 'LIMITED', 'LTD',
            'FARMERS', 'UNION', 'MPCS', 'FCSL', 'MPCSL',
            # Instructions and messages
            'YOU JUST HAVE', 'TILL', 'TOMORROW', 'SEND YOUR DETAILS',
            'WHATSAPP', 'SMS', 'DON\'T SEND', 'OTHER NUMBERS',
            'PLEASE', 'HELP', 'CORRECTION', 'THANK YOU', 'CORRECT',
            'SERIAL NO', 'INSTEAD OF', 'SPELLED', 'SIR',
            # Security/company names
            'PROTECTUS SECURITY', 'INTERNATIONAL',
            # Special characters and formatting
            '👆', '👇🏻', '👇', '🏻'
        ]
        
        # Check if line contains noise patterns
        for pattern in noise_patterns:
            if pattern in line_upper:
                return True
        
        # Check for lines that are all caps and look like titles (but not key-value pairs)
        if (line.isupper() and 
            len(line) > 20 and 
            not ':' in line and
            ('COOPERATIVE' in line or 'SOCIETY' in line or 'LIMITED' in line)):
            return True
        
        # Check for lines with excessive punctuation
        punct_count = sum(1 for c in line if c in '.,!?;')
        if len(line) > 10 and punct_count / len(line) > 0.3:
            return True
        
        # Check for lines that look like serial numbers or corrections
        if (line_upper.startswith('SERIAL') or 
            'CORRECT' in line_upper or 
            'SPELLING' in line_upper):
            return True
        
        return False
    
    def is_valid_key_value_pair(self, line):
        """
        Check if a line represents a valid KEY: VALUE pair for data records.
        
        Args:
            line (str): Line to check
            
        Returns:
            bool: True if it's a valid key-value pair
        """
        if ':' not in line:
            return False
        
        parts = line.split(':', 1)
        key = parts[0].strip().upper()
        value = parts[1].strip()
        
        # Valid key patterns (common data fields)
        valid_keys = [
            'NAME', 'PHONE', 'PHONE NO', 'BANK', 'BANK NAME', 'ACCT', 'ACCT NO', 
            'ACCOUNT', 'ACCOUNT NO', 'SEX', 'EMAIL', 'ADDRESS', 'CEO', 'CEO NAME',
            'CO-OP NAME', 'COOP NAME', 'COOPERATIVE NAME', 'COOPERATIVE', 'CO-OP',
            'MEMBER ID', 'ID', 'GENDER'
        ]
        
        # Check if key matches valid patterns
        for valid_key in valid_keys:
            if valid_key in key:
                return True
        
        # Additional checks for common variations
        if (any(word in key for word in ['NAME', 'PHONE', 'BANK', 'ACCT', 'SEX']) and 
            len(key) <= 30):  # Reasonable key length
            return True
        
        return False
    
    def is_record_end_field(self, line):
        """
        Check if a line contains the record end field.
        
        Args:
            line (str): Line to check
            
        Returns:
            bool: True if line contains the record end field
        """
        if not self.record_end_field:
            # Default to SEX if no end field configured, but be more conservative
            return self.is_conservative_sex_field(line)
        
        # Extract key from line if it's a key-value pair
        key, value = self.extract_key_value_from_line(line)
        if key:
            key_normalized = self.normalize_key_name(key)
            end_normalized = self.normalize_field_for_comparison(self.record_end_field)
            return key_normalized == end_normalized
        
        # For non-key-value lines, check if the line contains the field
        line_upper = line.upper()
        end_field_upper = self.record_end_field.upper()
        
        # Handle SEX field specifically
        if end_field_upper == "SEX":
            return self.is_conservative_sex_field(line)
        
        return end_field_upper in line_upper
    
    def is_conservative_sex_field(self, line):
        """
        More conservative SEX field detection that only triggers record end
        when we're confident it's actually the end of a complete record.
        
        Args:
            line (str): Line to check
            
        Returns:
            bool: True if line should end a record
        """
        line_upper = line.upper().strip()
        
        # Only explicit SEX field formats should end records
        if ('SEX:' in line_upper and any(gender in line_upper for gender in ['MALE', 'FEMALE'])) or \
           (line_upper.startswith('SEX') and any(gender in line_upper for gender in ['MALE', 'FEMALE'])):
            return True
        
        # Don't use standalone gender values as record boundaries
        # Let the intelligent orphaned value handler deal with them
        return False
    
    def is_record_start_field(self, line):
        """
        Check if a line contains the record start field.
        
        Args:
            line (str): Line to check
            
        Returns:
            bool: True if line contains the record start field
        """
        if not self.record_start_field:
            return False
        
        # Extract key from line if it's a key-value pair
        key, value = self.extract_key_value_from_line(line)
        if key:
            key_normalized = self.normalize_key_name(key)
            start_normalized = self.normalize_field_for_comparison(self.record_start_field)
            return key_normalized == start_normalized
        
        # For non-key-value lines, check if the line contains the field
        line_upper = line.upper()
        start_field_upper = self.record_start_field.upper()
        
        # Handle common variations
        if start_field_upper == "NAME OF COOPERATIVE":
            return any(pattern in line_upper for pattern in [
                "COOP", "CO-OP", "COOPERATIVE", "ORGANIZATION", "COMPANY"
            ]) and "NAME" in line_upper
        
        return start_field_upper in line_upper
    
    def is_sex_field(self, line):
        """
        Check if a line contains SEX information (fallback for record end).
        
        Args:
            line (str): Line to check
            
        Returns:
            bool: True if line contains SEX information that should end a record
        """
        line_upper = line.upper().strip()
        
        # Only consider it a record-ending SEX field if:
        # 1. It has explicit SEX: format, OR
        # 2. It's a standalone gender value AND it's at the end of a meaningful record block
        
        # Explicit SEX field formats
        if ('SEX:' in line_upper and any(gender in line_upper for gender in ['MALE', 'FEMALE'])) or \
           (line_upper.startswith('SEX') and any(gender in line_upper for gender in ['MALE', 'FEMALE'])):
            return True
        
        # For standalone gender values, only treat as record end if it seems appropriate
        # (this is less aggressive than before)
        if line_upper in ['MALE', 'FEMALE']:
            # Don't automatically treat standalone gender as record end
            # Let the intelligent orphaned value handler deal with this
            return False
        
        return False
    
    def normalize_field_for_comparison(self, field_name):
        """
        Normalize a field name for comparison with extracted keys.
        
        Args:
            field_name (str): Field name to normalize
            
        Returns:
            str: Normalized field name
        """
        field_upper = field_name.upper()
        
        # Map common field variations to normalized names
        field_mappings = {
            "NAME OF COOPERATIVE": "CO-OP NAME",
            "CEO NAME": "CEO NAME", 
            "PHONE No.": "PHONE NO",
            "BANK NAME": "BANK NAME",
            "ACNT. No.": "ACNT NO",
            "SEX": "SEX"
        }
        
        return field_mappings.get(field_upper, field_upper)

    def normalize_key_name(self, key):
        """
        Normalize a key name to standard format with enhanced field recognition.
        
        Args:
            key (str): Original key name
            
        Returns:
            str: Normalized key name
        """
        key_upper = key.upper().strip()
        
        # Remove common noise words and clean up (but preserve COOPERATIVE NAME pattern)
        key_upper = (key_upper.replace('PERSONAL ', '')
                              .replace("PERSONAL. ", "")
                              .replace('CEO\'S ', '')
                              .replace('.', ' ')
                              .replace('  ', ' ')
                              .strip())
        
        # Standardize common variations - order matters!
        # First handle explicit cooperative name patterns
        if (any(pattern in key_upper for pattern in [
            'COOPERATIVE NAME', 'CO-OPERATIVE NAME', 'COOP NAME', 'CO-OP NAME',
            'ORGANIZATION NAME', 'ORGANISATION NAME', 'COMPANY NAME'
        ]) or (any(word in key_upper for word in ['COOPERATIVE', 'CO-OPERATIVE', 'COOP', 'CO-OP', 'ORGANISATION', 'ORGANIZATION', 'NGO']) and 'NAME' in key_upper)):
            return 'CO-OP NAME'
        
        # Handle standalone cooperative identifiers (without "NAME")
        elif key_upper in ['COOPERATIVE', 'CO-OPERATIVE', 'COOP', 'CO-OP', 'ORGANIZATION', 'ORGANISATION', 'COMPANY', 'NGO']:
            return 'CO-OP NAME'
        
        # Handle various name fields (be more specific to avoid conflicts)
        elif (key_upper in ['CEO NAME', 'PERSONAL NAME', 'FULL NAME', 'CEO'] or 
              'CEO NAME' in key_upper or 'PERSONAL NAME' in key_upper or 'FULL NAME' in key_upper or
              ('CEO' in key_upper and 'NAME' in key_upper)):
            return 'CEO NAME'
        
        # Handle simple "NAME" field - but only if not already handled above
        elif key_upper == 'NAME':
            return 'CEO NAME'  # Default to CEO NAME for consistency
        
        # Phone number variations (enhanced)
        elif any(phrase in key_upper for phrase in [
            'PHONE NO', 'PHONE', 'GSM', 'MOBILE', 'TEL', 'TELEPHONE',
            'CONTACT', 'CELL', 'NUMBER'
        ]) and not any(block in key_upper for block in ['ACCOUNT', 'ACCT', 'BANK']):
            return 'PHONE NO'
        
        # Bank name variations (enhanced)
        elif any(phrase in key_upper for phrase in [
            'BANK NAME', 'BANK', 'FINANCIAL INSTITUTION'
        ]) and 'ACCOUNT' not in key_upper and 'ACCT' not in key_upper:
            return 'BANK NAME'
        
        # Account number variations (enhanced) - must come after bank name check
        elif any(phrase in key_upper for phrase in [
            'ACNT NO', 'ACCOUNT NO', 'ACC NO', 'ACCT NO', 'A/C NO',
            'ACCOUNT NUMBER', 'ACCT NUMBER', 'ACC NUMBER', 'A/C NUMBER',
            'ACCOUNT', 'ACCT', 'ACC', 'A/C'
        ]):
            return 'ACNT NO'
        
        # Gender/Sex variations (enhanced)
        elif key_upper in ['SEX', 'GENDER', 'M/F', 'MALE/FEMALE']:
            return 'SEX'
        
        # Standard fields
        elif key_upper in ['EMAIL', 'E-MAIL', 'MAIL']:
            return 'EMAIL'
        elif key_upper in ['ADDRESS', 'LOCATION', 'ADDR']:
            return 'ADDRESS'
        
        # If it just says BANK without specification, assume BANK NAME
        elif key_upper == 'BANK':
            return 'BANK NAME'
        
        # If it contains PHONE but we missed it somehow
        elif 'PHONE' in key_upper or 'GSM' in key_upper:
            return 'PHONE NO'
        
        else:
            return key_upper
    
    def clean_and_normalize_data(self, records, headers):
        """
        Clean and normalize the extracted data records.
        
        Args:
            records (list): List of record rows
            headers (list): List of column headers
            
        Returns:
            tuple: (cleaned_headers, cleaned_records)
        """
        if not records or not headers:
            return headers, records
        
        # Normalize headers
        normalized_headers = []
        for header in headers:
            normalized_headers.append(self.normalize_key_name(header))
        
        # Remove duplicates while preserving order
        seen = set()
        final_headers = []
        for header in normalized_headers:
            if header not in seen:
                seen.add(header)
                final_headers.append(header)
        
        # Clean records
        cleaned_records = []
        for record in records:
            cleaned_record = []
            for i, value in enumerate(record):
                if i < len(final_headers):
                    # Clean up values
                    clean_value = str(value).strip()
                    
                    # Remove common suffixes/prefixes
                    clean_value = clean_value.replace('Account Name:', '').strip()
                    clean_value = clean_value.replace('ACC No', '').strip()
                    clean_value = clean_value.replace('Acct. N0.', '').strip()
                    clean_value = clean_value.replace('Phone no.', '').strip()
                    clean_value = clean_value.replace('CEO:', '').strip()
                    
                    cleaned_record.append(clean_value)
                else:
                    cleaned_record.append(str(value).strip())
            
            # Ensure all records have the same number of columns
            while len(cleaned_record) < len(final_headers):
                cleaned_record.append('')
            
            # Only add records that have some actual data
            if any(val.strip() for val in cleaned_record):
                cleaned_records.append(cleaned_record[:len(final_headers)])
        
        return final_headers, cleaned_records
    
    def parse_records(self, text):
        """
        Parse records from text input and map to standard columns.
        
        Args:
            text (str): Raw text input containing records
            
        Returns:
            tuple: (column_headers, records) where records is list of lists with S/N generated
        """
        lines = [line.strip() for line in text.strip().split('\n') if line.strip()]
        
        # Split into record blocks using intelligent boundary detection
        record_blocks = []
        current_block = []
        
        for i, line in enumerate(lines):
            # Skip obvious noise lines
            if self.is_noise_line(line):
                continue
            
            # Check if this line is a record start field (indicates new record)
            if self.is_record_start_field(line) and current_block:
                # Save current block and start new one
                record_blocks.append(current_block[:])
                current_block = [line]
            # Check for intelligent record boundary detection
            elif self.is_intelligent_record_boundary(line, current_block, lines, i):
                # End current record and start new one
                if current_block:
                    record_blocks.append(current_block[:])
                current_block = [line]
            else:
                current_block.append(line)
            
            # Check if this line contains the record end field (but be more careful)
            if self.is_record_end_field(line) and current_block:
                # Only end record if we have substantial content
                if len(current_block) >= 2:  # At least 2 lines for a valid record
                    record_blocks.append(current_block[:])
                    current_block = []
        
        # Don't forget the last block if it doesn't end with the end field
        if current_block:
            record_blocks.append(current_block)
        
        if not record_blocks:
            return [], []
        
        # Post-process: merge orphaned value blocks with preceding complete records
        merged_blocks = self.merge_orphaned_blocks(record_blocks)
        
        # Process each record block and map to standard columns
        records_data = []
        
        for block in merged_blocks:
            record = self.extract_record_data(block)
            if record:
                records_data.append(record)
        
        if not records_data:
            return [], []
        
        # Generate standardized output using user-configured columns
        standardized_records = []
        serial_number = 1
        
        for record_dict in records_data:
            standardized_record = []
            
            for column in self.user_columns:
                if column == "S/N":
                    standardized_record.append(str(serial_number))
                else:
                    # Map the column to the record data
                    value = self.map_column_to_data(column, record_dict)
                    standardized_record.append(value)
            
            standardized_records.append(standardized_record)
            serial_number += 1
        
        return self.user_columns.copy(), standardized_records
    
    def extract_record_data(self, block):
        """Extract structured data from a record block with enhanced field detection."""
        record = {}
        block_text = '\n'.join(block)
        
        # Extract SEX field first (it's usually at the end)
        sex_value = self.extract_sex_from_block(block_text)
        if sex_value:
            record['SEX'] = self.clean_value(sex_value)
        
        # Process each line in the block
        i = 0
        while i < len(block):
            line = block[i].strip()
            
            # Try to parse as key-value pair (handles various separators and patterns)
            key, value = self.extract_key_value_from_line(line)
            if key and value:
                normalized_key = self.normalize_key_name(key)
                record[normalized_key] = self.clean_value(value)
            
            # Handle multiline fields (when value continues on next line)
            elif line and key is None:  # Only if no key-value pair was found
                # This might be a continuation of the previous field
                prev_line = block[i-1].strip()
                if any(sep in prev_line for sep in [':', '. ', '-']):
                    if any(prev_line.endswith(end) for end in [':', '.', '-']):
                        # Previous line was a key without value
                        key_part = prev_line.replace(':', '').replace('.', '').replace('-', '').strip()
                        # Handle special cases like "ACCOUNT NUMBER" or "PHONE NUMBER" 
                        key_upper = key_part.upper()
                        if any(field_word in key_upper for field_word in [
                            'ACCOUNT', 'ACCT', 'ACNT', 'PHONE', 'BANK', 'NAME', 'SEX'
                        ]):
                            normalized_key = self.normalize_key_name(key_part)
                            record[normalized_key] = self.clean_value(line)
                    else:
                        # Check if current line looks like a value for a field we recognize
                        self.try_match_orphaned_value(line, record, i, block)
            
            i += 1
        
        # Enhanced pattern matching for missing fields
        self.extract_enhanced_patterns(block_text, record)
        
        # Clean all values in the record
        for key in record:
            record[key] = self.clean_value(record[key])
        
        # Only return records that have meaningful data (at least 2 fields)
        return record if len(record) >= 2 else None
    
    def clean_value(self, value):
        """Clean a field value by removing unwanted characters and formatting."""
        if not value:
            return ""
        
        # Convert to string and strip whitespace
        clean_val = str(value).strip()
        
        # For phone numbers, keep only digits (but preserve formatting for readability)
        if len(clean_val) == 11 and clean_val.isdigit():
            return clean_val
        
        # For account numbers, keep only digits
        if any(keyword in clean_val.upper() for keyword in ['ACCOUNT', 'ACCT', 'A/C']):
            # Extract just the digits
            digits = ''.join(filter(str.isdigit, clean_val))
            return digits if len(digits) >= 8 else clean_val
        
        # Remove excessive whitespace and special formatting characters
        clean_val = ' '.join(clean_val.split())
        
        # Remove common unwanted characters but keep essential ones
        unwanted_chars = ['*', '"', "'", '`', '~', '#', '$', '%', '^', '&']
        for char in unwanted_chars:
            clean_val = clean_val.replace(char, '')
        
        # Clean up common suffixes/prefixes that appear in raw data
        prefixes_to_remove = ['Account Name:', 'ACC No:', 'Acct. N0.:', 'Phone no.:', 'CEO:', 'Name:']
        for prefix in prefixes_to_remove:
            if clean_val.startswith(prefix):
                clean_val = clean_val[len(prefix):].strip()
        
        # Remove trailing periods for simple values (like MALE/FEMALE)
        if clean_val.upper() in ['MALE.', 'FEMALE.', 'M.', 'F.']:
            clean_val = clean_val.rstrip('.')
        
        return clean_val.strip()
    
    def try_match_orphaned_value(self, line, record, current_index, block):
        """Try to match an orphaned value to a field based on intelligent context analysis."""
        line_upper = line.upper().strip()
        line_clean = line.strip()
        
        # Look at previous and next lines for context
        context_lines = []
        for j in range(max(0, current_index - 3), min(len(block), current_index + 2)):
            if j < len(block) and j != current_index:
                context_lines.append(block[j].upper().strip())
        
        context_text = ' '.join(context_lines)
        
        # PRIORITY 1: SEX value detection (high priority) - intelligent gender recognition
        if line_upper in ['MALE', 'FEMALE', 'M', 'F']:
            if 'SEX' not in record or not record['SEX']:
                record['SEX'] = line_clean.title()
                return True
            else:
                # SEX already set, don't use this value for anything else
                return True
        
        # PRIORITY 2: Intelligent bank name detection - handles standalone bank names
        bank_keywords = ['BANK', 'UBA', 'GTB', 'FIRST', 'ACCESS', 'ZENITH', 'UNION', 'FCMB', 
                        'FIDELITY', 'ECO', 'STERLING', 'WEMA', 'KEYSTONE', 'POLARIS', 'GUARANTY']
        
        # Enhanced bank detection logic
        is_likely_bank = False
        if any(keyword in line_upper for keyword in bank_keywords):
            # Direct bank name indicators
            if 'BANK' in line_upper or 'PLC' in line_upper:
                is_likely_bank = True
            # Known bank acronyms or names
            elif any(bank in line_upper for bank in ['UBA', 'GTB', 'FCMB', 'ACCESS', 'ZENITH', 'FIRST']):
                is_likely_bank = True
            # Context-based detection (previous lines mention bank)
            elif any(keyword in context_text for keyword in ['BANK', 'PERSONAL BANK']):
                is_likely_bank = True
        
        # Smart bank name assignment - prioritize if not already set or if current value seems wrong
        if is_likely_bank and line_upper not in ['MALE', 'FEMALE', 'M', 'F']:
            current_bank = record.get('BANK NAME', '')
            # Assign if no bank name exists, or replace if current one looks wrong
            if (not current_bank or 
                current_bank.upper() in ['MALE', 'FEMALE', 'M', 'F'] or
                len(current_bank) < 3):
                record['BANK NAME'] = line_clean
                return True
        
        # PRIORITY 3: Phone number detection (11 digits)
        if line_clean.isdigit() and len(line_clean) == 11:
            if (any(keyword in context_text for keyword in ['PHONE', 'GSM', 'MOBILE', 'TEL', 'CONTACT']) or
                'PHONE NO' not in record):
                if 'PHONE NO' not in record:
                    record['PHONE NO'] = line_clean
                return True
        
        # PRIORITY 4: Account number detection - enhanced patterns
        digits_only = ''.join(filter(str.isdigit, line))
        if len(digits_only) >= 8:
            # Check for account number context
            if any(keyword in context_text for keyword in ['ACCOUNT', 'ACCT', 'A/C']):
                if 'ACNT NO' not in record:
                    record['ACNT NO'] = digits_only
                return True
            # Also check previous line for account patterns
            if current_index > 0:
                prev_line = block[current_index - 1].upper()
                if any(pattern in prev_line for pattern in ['A/C NO', 'ACCOUNT', 'ACCT']):
                    if 'ACNT NO' not in record:
                        record['ACNT NO'] = digits_only
                    return True
        
        # PRIORITY 5: Person name continuation (multi-line names)
        if (len(line_clean.split()) <= 3 and 
            not line_clean.isdigit() and 
            current_index > 0 and
            len(line_clean) > 2 and
            line_upper not in ['MALE', 'FEMALE', 'M', 'F']):  # Exclude gender values
            
            prev_line = block[current_index - 1].upper()
            # Check if previous line was a CEO name field
            if any(pattern in prev_line for pattern in ['CEO NAME', 'PERSONAL NAME', 'NAME']):
                # This might be a continuation of the person's name
                if 'CEO NAME' in record:
                    # Append to existing CEO name
                    record['CEO NAME'] = (record['CEO NAME'] + ' ' + line_clean).strip()
                    return True
                elif 'CEO NAME' not in record:
                    # Store as CEO name if no specific field found yet
                    record['CEO NAME'] = line_clean
                    return True
        
        # PRIORITY 6: Cooperative name detection for orphaned organization names
        coop_indicators = ['MPCS', 'LTD', 'LIMITED', 'COOPERATIVE', 'SOCIETY', 'ENTERPRISE']
        if (any(word in line_upper for word in coop_indicators) and
            ('CO-OP NAME' not in record or not record['CO-OP NAME'])):
            record['CO-OP NAME'] = line_clean
            return True
        
        # PRIORITY 7: Intelligent field assignment based on missing fields and content analysis
        # If we have a standalone value that hasn't been assigned yet, try intelligent mapping
        if len(line_clean) > 1 and not line_clean.isdigit():
            # Check what fields are still missing and try to intelligently assign
            missing_fields = []
            if 'SEX' not in record or not record['SEX']:
                missing_fields.append('SEX')
            if 'BANK NAME' not in record or not record['BANK NAME'] or record.get('BANK NAME', '').upper() in ['MALE', 'FEMALE']:
                missing_fields.append('BANK NAME')
            if 'CEO NAME' not in record or not record['CEO NAME']:
                missing_fields.append('CEO NAME')
            if 'CO-OP NAME' not in record or not record['CO-OP NAME']:
                missing_fields.append('CO-OP NAME')
            
            # Smart assignment based on content characteristics
            if missing_fields:
                # Gender values go to SEX
                if line_upper in ['MALE', 'FEMALE', 'M', 'F'] and 'SEX' in missing_fields:
                    record['SEX'] = line_clean.title()
                    return True
                
                # Bank-like content goes to BANK NAME
                elif (any(keyword in line_upper for keyword in bank_keywords) and 
                      'BANK NAME' in missing_fields):
                    record['BANK NAME'] = line_clean
                    return True
                
                # Organization indicators go to CO-OP NAME
                elif (any(word in line_upper for word in coop_indicators) and 
                      'CO-OP NAME' in missing_fields):
                    record['CO-OP NAME'] = line_clean
                    return True
                
                # Person name-like content (2-3 words, title case) goes to CEO NAME
                elif (len(line_clean.split()) in [2, 3] and 
                      line_clean.istitle() and 
                      'CEO NAME' in missing_fields and
                      not any(word in line_upper for word in ['BANK', 'COOPERATIVE', 'SOCIETY', 'LIMITED', 'LTD'])):
                    record['CEO NAME'] = line_clean
                    return True
        
        return False
    
    def extract_enhanced_patterns(self, block_text, record):
        """Enhanced pattern matching for various field formats including multiline."""
        import re
        
        # Clean asterisks from text first for better pattern matching
        clean_block_text = block_text.replace('*', '')
        
        # Enhanced phone number patterns (including semicolon separators and GSM)
        phone_patterns = [
            r'(?:\*?(?:PHONE|GSM|MOBILE|TEL)\*?)(?:\s*NO\.?\*?|\s*NUMBER\*?)?(?:\s*[:.,-;]?\*?\s*)(\d{10,11})',
            r'(?:\*?(?:PHONE|GSM|MOBILE|TEL)\*?)(?:\s*NO\.?\*?|\s*NUMBER\*?)?(?:\s*[:.,-;]?\*?\s*\n\s*)(\d{10,11})',
            r'(?:PHONE\s*NUMBER)\s*\n\s*(\d{10,11})',  # Multiline "PHONE NUMBER \n 08028..."
            r'(\d{4}[-.\s]?\d{3}[-.\s]?\d{4})',  # Formatted phone patterns
            r'GSM:\s*(\d{10,11})',  # Direct GSM: format
            r'(?:^|\n)\s*(\d{11})\s*(?=\n|$)',  # Standalone 11-digit numbers
            r'(?:^|\n)\s*(\d{10})\s*(?=\n|$)',  # Standalone 10-digit numbers
        ]
        
        for pattern in phone_patterns:
            matches = re.findall(pattern, block_text, re.MULTILINE | re.IGNORECASE)
            if matches and not record.get('PHONE NO'):
                # Clean the match
                phone = ''.join(filter(str.isdigit, matches[0]))
                if len(phone) >= 10:  # Accept both 10 and 11 digit numbers
                    record['PHONE NO'] = phone
                    break
        
        # Enhanced account number patterns (including semicolon, A/C, ACT. NO, and missing spaces)
        account_patterns = [
            r'(?:\*?(?:ACCT?|ACCOUNT)\*?)(?:\s*(?:NO?|NUMBER)\*?\.?)?(?:\s*[:.,-;]?\*?\s*)\*?(\d{8,})\*?',
            r'(?:\*?(?:ACCT?|ACCOUNT)\*?)(?:\s*(?:NO?|NUMBER)\*?\.?)?(?:\s*[:.,-;]?\*?\s*\n\s*)\*?(\d{8,})\*?',
            r'(?:ACCOUNT\s*NUMBER)\s*\n\s*(\d{8,})',  # Multiline "ACCOUNT NUMBER \n 4091..."
            r'(?:A/C|ACC)(?:\s*(?:NO\.?|NUMBER))?(?:\s*[:.,-;]?\s*)(\d{8,})',
            r'(?:\*?PERSONAL\*?\.?\s*\*?(?:ACNT?|ACCOUNT)\*?)(?:\s*(?:NO?\*?\.?|NUMBER\*?)\s*)(\d{8,})',
            r'(?:PERSONAL\s*)?(?:ACNT?|ACCOUNT)\.?\s*NO?\.?\s*(\d{8,})',  # Flexible period format
            r'(?:ACT\.?\s*NO?\.?):\s*(\d{8,})',  # ACT. NO.: format
            r'(?:ACNT?\.\s*No\.)(\d{8,})',  # Missing space after period: "ACNT. No.1234"
            r'(?:ACNT?\.\s*N0\.?):\s*(\d{8,})',  # "ACNT. N0.:" format (typo in NO)
            r'Acc\.?\s*[Nn]o\.?\s*[:.]\s*(\d{8,})',  # "Acc No:" format
            r'(?:ACNT?\.\s*No)\s*\*?(\d{8,})\*?',  # "ACNT. No *1234*" format
            r'(?:^|\n)\s*(\d{10})\s*(?=\n|$)',  # Standalone 10-digit account numbers
            r'(?:^|\n)\s*(\d{8,9})\s*(?=\n|$)',  # Standalone 8-9 digit account numbers
        ]
        
        for pattern in account_patterns:
            matches = re.findall(pattern, block_text, re.MULTILINE | re.IGNORECASE)
            if matches and not record.get('ACNT NO'):
                # Clean the match - keep only digits
                account = ''.join(filter(str.isdigit, matches[0]))
                if len(account) >= 8:
                    record['ACNT NO'] = account
                    break
        
        # Enhanced bank name patterns (more conservative to avoid false matches)
        bank_patterns = [
            r'(?:BANK\s*NAME)(?:\s*[:.,-;]?\s*)([A-Z][A-Z\s&.]+?)(?=\n|$)',  # "BANK NAME: VALUE"
            r'(?:PERSONAL\s*BANK\s*NAME)(?:\s*[:.,-;]?\s*)([A-Z][A-Z\s&.]+?)(?=\n|$)',  # "PERSONAL BANK NAME: VALUE"
            r'(?:^|\n)\s*([A-Z]{2,}\s*BANK(?:\s+PLC)?)\s*(?=\n|$)',  # Standalone bank names like "ZENITH BANK"
            r'(?:^|\n)\s*(ACCESS\s*BANK|ZENITH\s*BANK|FIRST\s*BANK|UBA|GTB|UNION\s*BANK|POLARIS\s*BANK|FCMB|FIDELITY\s*BANK|STERLING\s*BANK)(?:\s*PLC)?\s*(?=\n|$)',  # Full bank names
            r'BANK:\s*([A-Z][A-Z\s&.]+?)(?=\n|$)',  # Direct "BANK: VALUE" format
        ]
        
        for pattern in bank_patterns:
            matches = re.findall(pattern, block_text, re.MULTILINE | re.IGNORECASE)
            if matches and not record.get('BANK NAME'):
                bank_name = matches[0].strip()
                # Additional validation to ensure this is actually a bank name
                if (len(bank_name) > 2 and 
                    not bank_name.isdigit() and 
                    bank_name.upper() not in ['MALE', 'FEMALE', 'M', 'F'] and  # Exclude gender values
                    ('BANK' in bank_name.upper() or any(bank in bank_name.upper() for bank in ['ACCESS', 'ZENITH', 'FIRST', 'UBA', 'GTB', 'POLARIS', 'FCMB', 'FIDELITY', 'STERLING']))):
                    record['BANK NAME'] = bank_name
                    break
        
        # Enhanced name patterns (more specific to avoid capturing bank names, etc.)
        name_patterns = [
            r'(?:\*?(?:CEO|PERSONAL)\*?)?\s*\*?NAME\*?(?:\s*[:.,-]?\*?\s*)([A-Z][A-Z\s.]+?)(?=\n|$)',  # CEO/PERSONAL NAME only
            r'(?:\*?(?:CEO|PERSONAL)\*?)?\s*\*?NAME\*?(?:\s*[:.,-]?\*?\s*\n\s*)([A-Z][A-Z\s.]+?)(?=\n|$)',
            r'(?:^|\n)(?:CEO\s*)?NAME\s*\n\s*([A-Z][A-Z\s.]+?)(?=\n|$)',  # Multiline "NAME \n PERSON NAME" but only if preceded by CEO
            r'CEO:\s*([A-Z][A-Z\s.]+?)(?=\n|$)',  # Direct CEO: format
            r'(?:FULL\s*NAME)(?:\s*[:.,-]?\s*)([A-Z][A-Z\s.]+?)(?=\n|$)',  # FULL NAME: format
            r'(?:PERSONAL\s*NAME)(?:\s*[:.,-]?\s*)([A-Z][A-Z\s.]+?)(?=\n|$)',  # PERSONAL NAME: format
            r'(?:NAME\s*OF\s*CEO|CEO\s*NAME):\s*\*?([A-Z][A-Z\s.]+?)\*?(?=\n|$)',  # "NAME OF CEO:" or "CEO NAME:"
            r'CEO\s*Name:\s*([A-Z][A-Z\s.]+?)(?=\n|$)',  # "CEO Name:" format
        ]
        
        for pattern in name_patterns:
            matches = re.findall(pattern, block_text, re.MULTILINE)
            if matches:
                name = matches[0].strip()
                if len(name) > 2 and not name.isdigit():
                    # Validate that this is actually a person's name, not a bank/org name
                    if not any(word in name.upper() for word in [
                        'BANK', 'ACCESS', 'ZENITH', 'FIRST', 'UBA', 'GTB', 'UNION', 'POLARIS',
                        'FCMB', 'FIDELITY', 'STERLING', 'COOPERATIVE', 'SOCIETY', 'COMPANY',
                        'ORGANIZATION', 'ENTERPRISE', 'LIMITED', 'LTD', 'PLC', 'COOP', 'NGO',
                        'PROBLEM'  # Also reject test data names
                    ]) and len(name.split()) <= 3:  # Person names usually 1-3 words
                        if not record.get('NAME'):
                            record['NAME'] = name
                        break
        
        # Enhanced cooperative name patterns (including ORGANIZATION, COMPANY, NAME OF CO-OPERATIVE)
        coop_patterns = [
            r'(?:\*?(?:COOPERATIVE|COOP|ORGANIZATION|COMPANY|NGO)\*?)(?:\s*NAME\*?)?(?:\s*[:.,-]?\*?\s*)([A-Z][A-Z\s&.()]+?)(?=\n|$)',
            r'(?:COOPERATIVE|COOP|ORGANIZATION|COMPANY|NGO)\.?\s+([A-Z][A-Z\s&.()]+?)(?=\n|$)',
            r'(?:NAME\s*OF\s*(?:CO-?OPERATIVE|COOPERATIVE)):\s*\*?([A-Z][A-Z\s&.()]+?)\*?(?=\n|$)',
            r'(?:NAME\s*OF\s*(?:CO-?OPERATIVE|COOPERATIVE)):\s*([A-Z][A-Z\s&.()]*\n[A-Z][A-Z\s&.()]*(?:SOCIETY|COOPERATIVE|MPCS)[A-Z\s.]*)',  # Multiline coop name
            r'(?:^|\n)\s*([A-Z][A-Z\s.()]*(?:MPCS|MPCSL|COOPERATIVE|SOCIETY|ENTERPRISE)[A-Z\s.]*)\s*(?=\n|$)',
            r'(?:CO-?OP\s*(?:NAME|name))(?:\s*[:.,-]?\s*)([A-Z][A-Z\s&.()]+?)(?=\n|$)',
            r'(?:CO-?OP\s*(?:NAME|name))\s*\n\s*([A-Z][A-Z\s&.()]+?)(?=\n|$)',  # Multiline CO-OP NAME
            r'(?:COOP\.\s*NAME):\s*([A-Z][A-Z\s&.()]+?)(?=\n|$)',  # "COOP. NAME:" format
            r'(?:COMPANY\s*NAME\.\s*-)\s*([A-Z][A-Z\s&.()]+?)(?=\n|$)',  # "COMPANY NAME. - VALUE" format
        ]
        
        for pattern in coop_patterns:
            matches = re.findall(pattern, block_text, re.MULTILINE | re.IGNORECASE)
            if matches and not record.get('COOPERATIVE'):
                coop_name = matches[0].strip()
                if len(coop_name) > 2 and not coop_name.isdigit():
                    record['COOPERATIVE'] = coop_name
                    break
        
        # Enhanced sex patterns (including missing colon and period after MALE/FEMALE)
        sex_patterns = [
            r'(?:\*?SEX\*?)(?:\s*[:.,-]?\*?\s*)\*?(MALE|FEMALE)\*?(?:\.|$|\s|\n)',
            r'SEX\s+\*?(MALE|FEMALE)\*?(?:\.|$|\s|\n)',  # "SEX MALE" format (missing colon)
        ]
        
        for pattern in sex_patterns:
            matches = re.findall(pattern, block_text, re.MULTILINE | re.IGNORECASE)
            if matches and not record.get('SEX'):
                sex = matches[0].strip().upper()
                if sex in ['MALE', 'FEMALE']:
                    record['SEX'] = sex
                    break
    
    def map_column_to_data(self, column, record_dict):
        """Map a standard column name to extracted record data."""
        column_upper = column.upper()
        
        # Direct mappings
        if column_upper == "NAME OF COOPERATIVE":
            # Try multiple variations of cooperative names
            for key in ['CO-OP NAME', 'COOP NAME', 'COOPERATIVE NAME', 'ORGANIZATION NAME']:
                if key in record_dict:
                    return record_dict[key]
            return ""
        
        elif column_upper == "CEO NAME":
            # Try multiple variations of name fields (prioritize specific fields)
            for key in ['CEO NAME', 'PERSONAL NAME']:
                if key in record_dict:
                    return record_dict[key]
            
            # Only use generic 'NAME' if it's clearly a person's name (not a bank/org name)
            if 'NAME' in record_dict:
                name_value = record_dict['NAME']
                # Reject if it looks like a bank name, organization name, or cooperative name
                if not any(word in name_value.upper() for word in [
                    'BANK', 'ACCESS', 'ZENITH', 'FIRST', 'UBA', 'GTB', 'UNION', 'POLARIS',
                    'FCMB', 'FIDELITY', 'STERLING', 'COOPERATIVE', 'SOCIETY', 'COMPANY',
                    'ORGANIZATION', 'ENTERPRISE', 'LIMITED', 'LTD', 'PLC', 'COOP', 'NGO',
                    'PROBLEM'  # Also reject test data names
                ]) and len(name_value.split()) <= 3:  # Person names usually 1-3 words
                    return name_value
            
            return ""
        
        elif column_upper == "PHONE NO.":
            # Try multiple variations of phone fields
            for key in ['PHONE NO', 'PERSONAL PHONE NO', 'PHONE NUMBER', 'PHONE']:
                if key in record_dict:
                    return record_dict[key]
            return ""
        
        elif column_upper == "BANK NAME":
            # Try multiple variations of bank fields
            for key in ['BANK NAME', 'PERSONAL BANK NAME', 'BANK']:
                if key in record_dict:
                    return record_dict[key]
            return ""
        
        elif column_upper == "ACNT. NO.":
            # Try multiple variations of account fields
            for key in ['ACNT NO', 'ACCOUNT NO', 'PERSONAL ACNT NO', 'ACC NO']:
                if key in record_dict:
                    return record_dict[key]
            return ""
        
        elif column_upper == "SEX":
            return record_dict.get('SEX', '')
        
        else:
            # For custom columns, try direct match first, then similar matches
            if column in record_dict:
                return record_dict[column]
            
            # Try case-insensitive match
            for key in record_dict:
                if key.upper() == column_upper:
                    return record_dict[key]
            
            return ""
    
    def extract_sex_from_block(self, block_text):
        """Extract SEX field value from a record block."""
        # Clean asterisks from text first
        text_clean = block_text.replace('*', '')
        text_upper = text_clean.upper()
        
        # Look for explicit SEX: patterns (enhanced with asterisk handling)
        import re
        sex_patterns = [
            r'SEX:\s*(MALE|FEMALE)',
            r'SEX\s*:?\s*(MALE|FEMALE)',
            r'SEX\s*[.:-]\s*(MALE|FEMALE)',
            r'\*SEX\*:\s*(MALE|FEMALE)',  # Asterisk format
            r'SEX\s*(MALE|FEMALE)',  # No colon format
        ]
        
        for pattern in sex_patterns:
            match = re.search(pattern, text_upper)
            if match:
                return match.group(1).title()
        
        # Look for standalone MALE/FEMALE at end of lines
        lines = text_clean.split('\n')
        for line in reversed(lines[-3:]):  # Check last 3 lines
            line_clean = line.strip().upper()
            if line_clean in ['MALE', 'FEMALE']:
                return line_clean.title()
        
        return None
    
    def extract_key_value_from_line(self, line):
        """Extract key and value from a line with enhanced separator handling."""
        key = None
        value = None
        
        # Clean asterisks from the line first (common formatting in user data)
        clean_line = line.replace('*', '').strip()
        
        # Handle special "A/C NO 1234567890" pattern (space-separated)
        import re
        account_match = re.match(r'(A/C\s*NO\.?)\s+(\d{8,})', clean_line, re.IGNORECASE)
        if account_match:
            key = account_match.group(1).strip()
            value = account_match.group(2).strip()
            return key, value
        
        # Handle other space-separated patterns like "PHONE NUMBER 08012345678"
        space_patterns = [
            r'(PHONE\s*NUMBER)\s+(\d{10,11})',
            r'(ACCOUNT\s*NUMBER)\s+(\d{8,})',
            r'(GSM\s*NO\.?)\s+(\d{10,11})',
            r'(MOBILE\s*NO\.?)\s+(\d{10,11})',
        ]
        
        for pattern in space_patterns:
            match = re.match(pattern, clean_line, re.IGNORECASE)
            if match:
                key = match.group(1).strip()
                value = match.group(2).strip()
                return key, value
        
        # Try colon separator first
        if ':' in clean_line:
            parts = clean_line.split(':', 1)
            if len(parts) == 2:
                key = parts[0].strip()
                value = parts[1].strip()
        
        # Try semicolon separator
        elif ';' in clean_line:
            parts = clean_line.split(';', 1)
            if len(parts) == 2:
                potential_key = parts[0].strip()
                potential_value = parts[1].strip()
                
                # Check if this looks like a field name
                key_upper = potential_key.upper()
                if any(field_word in key_upper for field_word in [
                    'NAME', 'PHONE', 'GSM', 'MOBILE', 'BANK', 'ACCOUNT', 'ACCT', 'ACNT', 'SEX', 'GENDER', 'NO', 'COOP'
                ]):
                    key = potential_key
                    value = potential_value
        
        # Try dash separator (e.g., "Phone no-08068066616")
        elif '-' in clean_line:
            parts = clean_line.split('-', 1)
            if len(parts) == 2:
                potential_key = parts[0].strip()
                potential_value = parts[1].strip()
                
                # Check if this looks like a field name
                key_upper = potential_key.upper()
                if any(field_word in key_upper for field_word in [
                    'NAME', 'PHONE', 'GSM', 'MOBILE', 'BANK', 'ACCOUNT', 'ACCT', 'ACNT', 'SEX', 'GENDER', 'NO', 'COOP'
                ]):
                    key = potential_key
                    value = potential_value
        
        # If no colon or dash, try period followed by space
        elif '. ' in clean_line:
            parts = clean_line.split('. ')
            
            if len(parts) == 2:
                # Standard case: "FIELD. VALUE"
                potential_key = parts[0].strip()
                potential_value = parts[1].strip()
                
                # Handle special case where the split gives us incomplete key/value
                if potential_value.startswith(('NO.', 'NO ', 'NUMBER')):
                    if 'NO.' in potential_value:
                        value_parts = potential_value.split('NO.', 1)
                        if len(value_parts) == 2:
                            potential_key = potential_key + " NO"
                            potential_value = value_parts[1].strip()
                    elif 'NO ' in potential_value:
                        value_parts = potential_value.split('NO ', 1)
                        if len(value_parts) == 2:
                            potential_key = potential_key + " NO"
                            potential_value = value_parts[1].strip()
                
                # Check if this looks like a field name
                key_upper = potential_key.upper()
                if any(field_word in key_upper for field_word in [
                    'NAME', 'PHONE', 'GSM', 'MOBILE', 'BANK', 'ACCOUNT', 'ACCT', 'ACNT', 'SEX', 'GENDER', 'NO'
                ]):
                    key = potential_key
                    value = potential_value
            
            elif len(parts) >= 3:
                # Multiple periods case
                potential_key_parts = []
                potential_value = None
                
                # Strategy 1: Look for numeric values
                for i, part in enumerate(parts):
                    part = part.strip()
                    if part.isdigit() and len(part) >= 8:  # Account numbers
                        potential_value = part
                        potential_key_parts = parts[:i]
                        break
                    elif part.isdigit() and len(part) == 11:  # Phone numbers
                        potential_value = part
                        potential_key_parts = parts[:i]
                        break
                
                # Strategy 2: Treat last part as value if it looks like a name/bank
                if not potential_value and len(parts) >= 2:
                    last_part = parts[-1].strip()
                    if (len(last_part) > 2 and 
                        not last_part.upper() in ['NO', 'NAME', 'PHONE', 'BANK', 'ACCOUNT'] and
                        any(char.isalpha() for char in last_part)):
                        potential_value = last_part
                        potential_key_parts = parts[:-1]
                
                if potential_value and potential_key_parts:
                    potential_key = ' '.join(potential_key_parts).strip()
                    key_upper = potential_key.upper()
                    if any(field_word in key_upper for field_word in [
                        'NAME', 'PHONE', 'GSM', 'MOBILE', 'BANK', 'ACCOUNT', 'ACCT', 'ACNT', 'SEX', 'GENDER', 'NO'
                    ]):
                        key = potential_key
                        value = potential_value
        
        # Try period with no space (e.g., "No.1023603764")
        elif '.' in clean_line and not '. ' in clean_line:
            # Look for pattern like "FIELD.VALUE" 
            parts = clean_line.split('.', 1)
            if len(parts) == 2:
                potential_key = parts[0].strip()
                potential_value = parts[1].strip()
                
                # Check if value starts with letters (indicating it's part of the field name)
                if potential_value and potential_value[0].isalpha():
                    # This is probably "FIELD.SUBFIELDVALUE" - try to extract numeric part
                    import re
                    numeric_match = re.search(r'(\d{8,})', potential_value)
                    if numeric_match:
                        # Found a long number - treat as account number
                        potential_key = potential_key + " " + re.sub(r'\d+', '', potential_value).strip()
                        potential_value = numeric_match.group(1)
                
                # Check if this looks like a field name
                key_upper = potential_key.upper()
                if any(field_word in key_upper for field_word in [
                    'NAME', 'PHONE', 'GSM', 'MOBILE', 'BANK', 'ACCOUNT', 'ACCT', 'ACNT', 'SEX', 'GENDER', 'NO'
                ]) and potential_value:
                    key = potential_key
                    value = potential_value
        
        # If still no key-value found, return None
        if not key:
            return None, None
        
        # Validate key (should be reasonable field name)
        if len(key) < 2 or len(key) > 50:
            return None, None
        
        # Skip if key looks like noise
        key_upper = key.upper()
        noise_keywords = ['YOU JUST', 'SEND YOUR', 'TILL 3PM', 'TOMORROW', "DON'T SEND"]
        if any(noise in key_upper for noise in noise_keywords):
            return None, None
        
        # If value is empty, we'll handle it in the main parsing logic
        return key, value if value else None
    
    def extract_patterns_from_block(self, block_text, record, all_keys=None):
        """Extract additional data using pattern matching."""
        import re
        
        # Common patterns for phone numbers
        phone_patterns = [
            r'(?:PHONE|Phone)(?:\s*No\.?|\s*NUMBER)?:?\s*(\d{11})',
            r'(?:GSM|Mobile):?\s*(\d{11})',
            r'(\d{4}[-.\s]?\d{3}[-.\s]?\d{4})',  # Phone format patterns
        ]
        
        for pattern in phone_patterns:
            match = re.search(pattern, block_text)
            if match and 'PHONE NO' not in record:
                record['PHONE NO'] = match.group(1)
                if all_keys is not None:
                    all_keys.add('PHONE NO')
                break
        
        # Account number patterns
        account_patterns = [
            r'(?:ACCT?|ACCOUNT)(?:\s*NO\.?|\s*NUMBER)?:?\s*(\d{10,})',
            r'A/C\s*(?:NO\.?|NUMBER)?:?\s*(\d{10,})',
        ]
        
        for pattern in account_patterns:
            match = re.search(pattern, block_text, re.IGNORECASE)
            if match and 'ACNT NO' not in record:
                record['ACNT NO'] = match.group(1)
                if all_keys is not None:
                    all_keys.add('ACNT NO')
                break
        
        # Bank name patterns
        bank_patterns = [
            r'BANK(?:\s*NAME)?:?\s*([A-Z][A-Z\s&]+?)(?:\n|$)',
            r'(?:^|\n)([A-Z]{2,}\s*BANK)(?:\s|$)',
        ]
        
        for pattern in bank_patterns:
            match = re.search(pattern, block_text, re.MULTILINE)
            if match and 'BANK NAME' not in record:
                bank_name = match.group(1).strip()
                if len(bank_name) > 2:
                    record['BANK NAME'] = bank_name
                    if all_keys is not None:
                        all_keys.add('BANK NAME')
                    break
    
    def group_by_coop_name(self, column_headers, records):
        """
        Group records by CO-OP NAME field.
        
        Args:
            column_headers (list): List of column header names
            records (list): List of record rows (each row is a list of values)
            
        Returns:
            dict: Dictionary with CO-OP NAME as keys and list of records as values
        """
        grouped = defaultdict(list)
        
        # Find the index of COOPERATIVE NAME column (using standard name)
        coop_name_index = None
        if "NAME OF COOPERATIVE" in column_headers:
            coop_name_index = column_headers.index("NAME OF COOPERATIVE")
        
        for record in records:
            if coop_name_index is not None and coop_name_index < len(record):
                coop_name = record[coop_name_index] if record[coop_name_index] else 'Unknown'
            else:
                coop_name = 'Unknown'
            grouped[coop_name].append(record)
        
        return grouped
    
    def create_excel_file(self, column_headers, grouped_data, filepath):
        """
        Create an Excel file with separate sheets for each CO-OP NAME.
        
        Args:
            column_headers (list): List of column header names
            grouped_data (dict): Grouped records by CO-OP NAME
            filepath (str): Path where Excel file should be saved
        """
        workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        for coop_name, records in grouped_data.items():
            if not records:
                continue
            
            # Create sheet name (Excel has 31 char limit and no invalid chars)
            sheet_name = coop_name[:31] if coop_name else "Unknown"
            
            # Remove invalid characters for Excel sheet names
            invalid_chars = ['*', '?', ':', '[', ']', '/', '\\']
            for char in invalid_chars:
                sheet_name = sheet_name.replace(char, '_')
            
            # Ensure sheet name is not empty and doesn't start/end with quotes
            sheet_name = sheet_name.strip().strip('\'"') or "Unknown"
            
            # Ensure unique sheet name
            original_name = sheet_name
            counter = 1
            while sheet_name in workbook.sheetnames:
                suffix = f"_{counter}"
                max_len = 31 - len(suffix)
                sheet_name = original_name[:max_len] + suffix
                counter += 1
            
            sheet = workbook.create_sheet(title=sheet_name)
            
            # Write header row
            for col_idx, header in enumerate(column_headers, start=1):
                sheet.cell(row=1, column=col_idx, value=header)
            
            # Write data rows
            for row_idx, record in enumerate(records, start=2):
                for col_idx, value in enumerate(record, start=1):
                    if col_idx <= len(column_headers):  # Don't exceed column count
                        sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(column_headers, start=1):
                max_length = len(str(header))
                for record in records:
                    if col_idx - 1 < len(record):
                        value = str(record[col_idx - 1])
                        max_length = max(max_length, len(value))
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Save the workbook
        workbook.save(filepath)
    
    def create_single_sheet_excel(self, column_headers, all_records, filepath):
        """
        Create an Excel file with all records in a single sheet.
        
        Args:
            column_headers (list): List of column header names
            all_records (list): All records as a flat list
            filepath (str): Path where Excel file should be saved
        """
        workbook = Workbook()
        
        # Use the default sheet or create one
        if 'Sheet' in workbook.sheetnames:
            sheet = workbook['Sheet']
            sheet.title = "All Records"
        else:
            sheet = workbook.create_sheet(title="All Records")
        
        # Write header row
        for col_idx, header in enumerate(column_headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            # Make headers bold
            cell.font = Font(bold=True)
        
        # Write data rows
        for row_idx, record in enumerate(all_records, start=2):
            for col_idx, value in enumerate(record, start=1):
                if col_idx <= len(column_headers):  # Don't exceed column count
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(column_headers, start=1):
            max_length = len(str(header))
            for record in all_records:
                if col_idx - 1 < len(record):
                    value = str(record[col_idx - 1])
                    max_length = max(max_length, len(value))
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Save the workbook
        workbook.save(filepath)
    
    def process_data(self):
        """Process the input data and export to Excel."""
        # Get text from text area
        text = self.text_area.get("1.0", tk.END)
        
        if not text.strip():
            messagebox.showwarning("No Data", "Please paste some data into the text area.")
            return
        
        # Parse records
        try:
            column_headers, records = self.parse_records(text)
            
            if not records:
                messagebox.showwarning("No Records", "No valid records found. Please check the format.")
                return
            
            if not column_headers:
                messagebox.showwarning("No Headers", "No column headers found. First record must use KEY: VALUE format.")
                return
            
            # Check user's choice for sheet organization
            sheet_mode = self.sheet_mode_var.get()
            
            # Ask user for save location
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Excel File As"
            )
            
            if not filepath:
                # User cancelled
                return
            
            # Create Excel file based on user choice
            if sheet_mode == "single":
                # Single sheet - all records together
                self.create_single_sheet_excel(column_headers, records, filepath)
                
                # Show confirmation
                messagebox.showinfo(
                    "Success",
                    f"Data successfully exported to single sheet!\n\nFile saved to:\n{filepath}\n\n"
                    f"Total records: {len(records)}\n"
                    f"Columns: {len(column_headers)}"
                )
            else:
                # Separate sheets by CO-OP NAME (default)
                grouped_data = self.group_by_coop_name(column_headers, records)
                
                if not grouped_data:
                    messagebox.showwarning("No Data", "No records to process.")
                    return
                
                self.create_excel_file(column_headers, grouped_data, filepath)
                
                # Show confirmation
                messagebox.showinfo(
                    "Success",
                    f"Data successfully exported to separate sheets!\n\nFile saved to:\n{filepath}\n\n"
                    f"Created {len(grouped_data)} sheet(s) with {len(records)} total record(s).\n"
                    f"Groups: {', '.join(grouped_data.keys())}"
                )
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")


def main():
    """Main entry point for the application."""
    root = tk.Tk()
    app = SmartDataExtractorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
