#!/usr/bin/env python3
"""
Simplified Data Sorter Application - Windows 7 Compatible Version
Minimal dependencies to avoid pywin32 and DLL loading issues
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import os
import sys
from pathlib import Path

# Import openpyxl components individually to avoid dependency issues
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    # Set dummy values to avoid unbound variable errors
    Workbook = None
    Font = None
    PatternFill = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")

class SimpleDataSorterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Sorter Application - Windows 7 Compatible")
        self.root.geometry("800x700")
        
        # Make window resizable
        self.root.rowconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        
        self.setup_ui()
        
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            messagebox.showwarning(
                "Missing Dependency", 
                "Excel export functionality is not available.\nText processing will still work."
            )
    
    def setup_ui(self):
        """Setup the user interface."""
        
        # Instructions
        instructions = tk.Label(
            self.root,
            text="Paste messy data below. App will filter noise and extract records automatically.",
            font=('Arial', 10, 'bold'),
            fg='#2c3e50',
            pady=10
        )
        instructions.grid(row=0, column=0, sticky='ew', padx=10)
        
        # Text input area with scrollbar
        self.text_area = scrolledtext.ScrolledText(
            self.root,
            wrap=tk.WORD,
            width=80,
            height=25,
            font=('Consolas', 9),
            bg='#f8f9fa',
            fg='#2c3e50'
        )
        self.text_area.grid(row=1, column=0, sticky='nsew', padx=10, pady=5)
        
        # Button frame
        button_frame = tk.Frame(self.root)
        button_frame.grid(row=2, column=0, pady=10)
        
        # Process button
        if OPENPYXL_AVAILABLE:
            process_btn_text = "Process and Export to Excel"
            process_btn_command = self.process_and_export
        else:
            process_btn_text = "Process Data (Text Output Only)"
            process_btn_command = self.process_text_only
        
        self.process_btn = tk.Button(
            button_frame,
            text=process_btn_text,
            command=process_btn_command,
            bg='#27ae60',
            fg='white',
            font=('Arial', 11, 'bold'),
            pady=8,
            padx=20
        )
        self.process_btn.pack(side=tk.LEFT, padx=5)
        
        # Clear button
        clear_btn = tk.Button(
            button_frame,
            text="Clear Text",
            command=self.clear_text,
            bg='#e74c3c',
            fg='white',
            font=('Arial', 11, 'bold'),
            pady=8,
            padx=20
        )
        clear_btn.pack(side=tk.LEFT, padx=5)
        
        # Example button
        example_btn = tk.Button(
            button_frame,
            text="Load Example",
            command=self.load_example,
            bg='#3498db',
            fg='white',
            font=('Arial', 11, 'bold'),
            pady=8,
            padx=20
        )
        example_btn.pack(side=tk.LEFT, padx=5)
        
        # Status label
        self.status_label = tk.Label(
            self.root,
            text="Ready to process data...",
            font=('Arial', 9),
            fg='#7f8c8d'
        )
        self.status_label.grid(row=3, column=0, pady=5)
    
    def is_noise_line(self, line):
        """Check if a line is noise/irrelevant."""
        line_clean = line.strip().upper()
        
        # Skip empty lines
        if not line_clean:
            return True
        
        # Common noise patterns
        noise_patterns = [
            'PERSONAL DATA', 'COOPERATIVE OWNERS', 'SEND YOUR DETAILS',
            'TILL 3PM', 'TOMORROW', "DON'T SEND", 'OTHER NUMBERS',
            'INSTRUCTIONS:', 'NOTE:', 'PLEASE', 'KINDLY'
        ]
        
        for pattern in noise_patterns:
            if pattern in line_clean:
                return True
        
        # Lines that are too short or too long are likely noise
        if len(line_clean) < 3 or len(line_clean) > 200:
            return True
        
        return False
    
    def is_valid_key_value_pair(self, line):
        """Check if line contains a valid key-value pair."""
        if ':' not in line:
            return False
        
        parts = line.split(':', 1)
        if len(parts) != 2:
            return False
        
        key = parts[0].strip()
        value = parts[1].strip()
        
        # Key should be reasonable length and value should exist
        return 2 <= len(key) <= 50 and len(value) > 0
    
    def normalize_key_name(self, key):
        """Normalize field names for consistency."""
        key = key.strip().upper()
        
        # Normalize common variations
        key_mappings = {
            'CEO NAME': 'NAME',
            'CEO': 'NAME',
            'PHONE NO': 'PHONE',
            'PHONE NUMBER': 'PHONE',
            'BANK NAME': 'BANK',
            'ACCT NO': 'ACCOUNT',
            'ACCOUNT NO': 'ACCOUNT',
            'ACC NO': 'ACCOUNT',
            'CO-OP NAME': 'COOP',
            'COOP NAME': 'COOP',
            'COOPERATIVE NAME': 'COOP'
        }
        
        return key_mappings.get(key, key)
    
    def parse_records(self, text):
        """Parse records from text input."""
        lines = text.strip().split('\n')
        records = []
        current_record = {}
        column_names = None
        
        self.status_label.config(text="Parsing data...")
        self.root.update()
        
        for line in lines:
            line = line.strip()
            
            # Skip noise lines
            if self.is_noise_line(line):
                continue
            
            # Empty line indicates record boundary
            if not line:
                if current_record and len(current_record) >= 2:  # At least 2 fields
                    records.append(current_record.copy())
                current_record = {}
                continue
            
            # Try to parse as key-value pair
            if self.is_valid_key_value_pair(line):
                key, value = line.split(':', 1)
                key = self.normalize_key_name(key)
                current_record[key] = value.strip()
                
                # Use first record to establish column names
                if column_names is None and current_record:
                    column_names = list(current_record.keys())
        
        # Don't forget the last record
        if current_record and len(current_record) >= 2:
            records.append(current_record)
        
        return records, column_names or []
    
    def group_by_coop(self, records):
        """Group records by cooperative name."""
        groups = {}
        
        for record in records:
            # Look for cooperative name with various field names
            coop_name = record.get('COOP', record.get('CO-OP NAME', record.get('COOPERATIVE', 'Unknown')))
            
            if coop_name not in groups:
                groups[coop_name] = []
            groups[coop_name].append(record)
        
        return groups
    
    def process_text_only(self):
        """Process data and show results in text format."""
        text_input = self.text_area.get("1.0", tk.END)
        
        if not text_input.strip():
            messagebox.showwarning("No Data", "Please enter some data to process.")
            return
        
        try:
            # Parse records
            records, column_names = self.parse_records(text_input)
            
            if not records:
                messagebox.showwarning("No Records", "No valid records found in the input data.")
                return
            
            # Group by cooperative
            groups = self.group_by_coop(records)
            
            # Create text output
            output_text = f"PROCESSED DATA SUMMARY\n{'=' * 50}\n\n"
            output_text += f"Total Records Found: {len(records)}\n"
            output_text += f"Number of Cooperatives: {len(groups)}\n"
            output_text += f"Columns: {', '.join(column_names)}\n\n"
            
            for coop_name, coop_records in groups.items():
                output_text += f"\n{coop_name.upper()}\n{'-' * len(coop_name)}\n"
                for i, record in enumerate(coop_records, 1):
                    output_text += f"\nRecord {i}:\n"
                    for col in column_names:
                        value = record.get(col, 'N/A')
                        output_text += f"  {col}: {value}\n"
            
            # Show in new window
            self.show_text_output(output_text)
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while processing data:\n{str(e)}")
    
    def show_text_output(self, text):
        """Show text output in a new window."""
        output_window = tk.Toplevel(self.root)
        output_window.title("Processed Data Output")
        output_window.geometry("600x500")
        
        text_widget = scrolledtext.ScrolledText(
            output_window,
            wrap=tk.WORD,
            font=('Consolas', 9)
        )
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_widget.insert("1.0", text)
        text_widget.config(state=tk.DISABLED)  # Make read-only
    
    def process_and_export(self):
        """Process data and export to Excel."""
        if not OPENPYXL_AVAILABLE:
            self.process_text_only()
            return
        
        text_input = self.text_area.get("1.0", tk.END)
        
        if not text_input.strip():
            messagebox.showwarning("No Data", "Please enter some data to process.")
            return
        
        try:
            # Parse records
            self.status_label.config(text="Parsing records...")
            self.root.update()
            
            records, column_names = self.parse_records(text_input)
            
            if not records:
                messagebox.showwarning("No Records", "No valid records found in the input data.")
                return
            
            # Group by cooperative
            self.status_label.config(text="Grouping by cooperative...")
            self.root.update()
            
            groups = self.group_by_coop(records)
            
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Excel file as..."
            )
            
            if not file_path:
                self.status_label.config(text="Export cancelled.")
                return
            
            # Create Excel file
            self.status_label.config(text="Creating Excel file...")
            self.root.update()
            
            # Check if openpyxl components are available
            if Workbook is None or Font is None or PatternFill is None or get_column_letter is None:
                raise ImportError("openpyxl components not properly loaded")
            
            workbook = Workbook()
            
            # Remove default worksheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Create sheets for each cooperative
            for coop_name, coop_records in groups.items():
                # Create safe sheet name
                safe_name = ''.join(c for c in coop_name if c.isalnum() or c in ' -_')[:31]
                worksheet = workbook.create_sheet(title=safe_name)
                
                # Add headers
                for col_idx, column_name in enumerate(column_names, 1):
                    cell = worksheet.cell(row=1, column=col_idx, value=column_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Add data
                for row_idx, record in enumerate(coop_records, 2):
                    for col_idx, column_name in enumerate(column_names, 1):
                        value = record.get(column_name, '')
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the file
            workbook.save(file_path)
            
            # Success message
            total_records = len(records)
            num_coops = len(groups)
            
            success_msg = f"✅ Export completed successfully!\n\n"
            success_msg += f"📁 File saved: {file_path}\n"
            success_msg += f"📊 {total_records} records processed\n"
            success_msg += f"🏢 {num_coops} cooperative groups created\n"
            success_msg += f"📋 {len(column_names)} columns: {', '.join(column_names)}"
            
            messagebox.showinfo("Export Successful", success_msg)
            self.status_label.config(text=f"Export completed: {total_records} records, {num_coops} groups")
            
        except ImportError as e:
            messagebox.showerror("Missing Dependencies", 
                               "Excel export functionality is not available.\n"
                               "The required 'openpyxl' library is not properly installed.\n\n"
                               "Using text output instead...")
            self.process_text_only()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
            self.status_label.config(text="Error occurred during processing")
    
    def clear_text(self):
        """Clear the text area."""
        self.text_area.delete("1.0", tk.END)
        self.status_label.config(text="Text cleared. Ready for new data.")
    
    def load_example(self):
        """Load example data."""
        example_data = '''PERSONAL DATA OF COOPERATIVE OWNERS

NAME: John Doe
CO-OP NAME: Alpha Co-op
PHONE NO: 08012345678
BANK NAME: First Bank
ACCT NO: 1234567890
SEX: MALE

YOU JUST HAVE NOW TILL 3PM TOMORROW TO SEND YOUR DETAILS
PLZ DON'T SEND TO OTHER NUMBERS

CEO NAME: Jane Smith
CO-OP NAME: Beta Co-op
PHONE NO: 08087654321
BANK NAME: GTB
ACCT NO: 0987654321
SEX: FEMALE

NAME: Bob Johnson
COOP NAME: Alpha Co-op
PHONE: 08055555555
BANK: UBA
ACCOUNT NO: 5555555555
SEX: MALE'''
        
        self.text_area.delete("1.0", tk.END)
        self.text_area.insert("1.0", example_data)
        self.status_label.config(text="Example data loaded. Click 'Process and Export' to test.")

def main():
    """Main function to run the application."""
    try:
        root = tk.Tk()
        app = SimpleDataSorterApp(root)
        root.mainloop()
    except Exception as e:
        # Fallback error handling
        try:
            import tkinter.messagebox as msgbox
            msgbox.showerror("Application Error", f"Failed to start application:\n{str(e)}")
        except:
            print(f"Failed to start application: {e}")
            input("Press Enter to exit...")

if __name__ == "__main__":
    main()